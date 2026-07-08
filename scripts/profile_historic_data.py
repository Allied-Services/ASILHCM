#!/usr/bin/env python3
"""Profile historic BPO/FM data for HCM migration (Build 1). Read-only."""
from __future__ import annotations
import csv, json, re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen
from openpyxl import load_workbook
REPO = Path(__file__).resolve().parent.parent
AUTODATA = Path(r"G:\My Drive\BPO & FM Business\AutoData from GSheets")
APRIL_2026 = Path(r"G:\My Drive\BPO & FM Business\BPOFM Payroll April 2026")
AUDIT_OUT = REPO / "audit" / "out"
SAMPLE_FILES = REPO / "frontend" / "public" / "SampleFiles"
AUDIT_ATTACH = REPO / "Attachments" / "Audit BPO FM Jul 2026"
PAYROLL_XLSX = AUDIT_ATTACH / "BPO FM Payroll & Invoice File.xlsx"
PR_CSV = REPO / "frontend" / "public" / "BPO FM Payroll & Invoice File - PR.csv"
LIVE_API = "https://asilhcm.onrender.com"
JWT_PATH = Path(r"C:\temp\hcm_jwt.txt")
MAP_PATH = REPO / "audit" / "data_migration_map.md"
RAW_JSON = REPO / "audit" / "profile_historic_data_raw.json"
MONTHLY_FOLDERS = ["2024_12", "2025_1", "2025_7", "2025_8", "2025_9"]
SENSITIVE_RE = re.compile(r"(cnic|account|iban|beneficiary\s*account|customer\s*reference|contact\s*number)", re.I)
DIGITS_RE = re.compile(r"\d")
_CACHE = {}

def mask_value(col, value):
    if value is None: return None
    s = str(value).strip()
    if not s: return s
    if SENSITIVE_RE.search(col or "") or len(DIGITS_RE.findall(s)) >= 8:
        if len(DIGITS_RE.findall(s)) >= 4: return re.sub(r"\d", "*", s[:-4]) + s[-4:]
    return s

def mask_row(columns, row):
    return {columns[i]: mask_value(columns[i], row[i] if i < len(row) else None) for i in range(len(columns))}

def detect_header_row(rows):
    best_idx, best_score = 0, -1
    for idx, row in enumerate(rows[:30]):
        texts = [str(c).strip() if c is not None else "" for c in row]
        non_empty = [t for t in texts if t]
        if len(non_empty) < 2: continue
        score = len(non_empty) + sum(1 for t in non_empty if len(t) > 2)
        if score > best_score: best_score, best_idx = score, idx
    return best_idx

def profile_gsheet(path):
    size = path.stat().st_size
    note = "Google-native .gsheet shortcut; unreadable locally."
    if size < 500: note += f" Stub size {size}B."
    return {"path": str(path), "kind": "gsheet", "readable": False, "note": note}

def profile_csv(path):
    key = str(path)
    if key in _CACHE: return _CACHE[key]
    rows = None
    for enc in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            with path.open("r", encoding=enc, newline="") as f: rows = list(csv.reader(f))
            break
        except Exception: continue
    if rows is None: prof = {"path": key, "kind": "csv", "error": "unreadable"}; _CACHE[key]=prof; return prof
    header_idx = detect_header_row([tuple(r) for r in rows[:15]])
    header = [str(c).strip() for c in rows[header_idx]]
    data_rows = [r for r in rows[header_idx+1:] if any(str(c).strip() for c in r)]
    prof = {"path": key, "kind": "csv", "header_row_index": header_idx, "columns": header,
            "row_count": len(data_rows), "sample_rows": [mask_row(header,r) for r in data_rows[:2]]}
    _CACHE[key]=prof; return prof

def profile_xlsx(path):
    key = str(path)
    if key in _CACHE: return _CACHE[key]
    prof = {"path": key, "kind": "xlsx", "sheets": []}
    try: wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e: prof["error"]=str(e); _CACHE[key]=prof; return prof
    for name in wb.sheetnames:
        ws = wb[name]; it = ws.iter_rows(values_only=True); first=[]
        for _ in range(25):
            try: first.append(next(it))
            except StopIteration: break
        hi = detect_header_row(first) if first else 0
        header = [str(c).strip() if c is not None else "" for c in first[hi]] if first else []
        count=0
        for row in it:
            if any(c is not None and str(c).strip() for c in row): count+=1
        for row in first[hi+1:]:
            if any(c is not None and str(c).strip() for c in row): count+=1
        samples=[mask_row(header,list(row)) for row in first[hi+1:hi+3]]
        prof["sheets"].append({"name":name,"header_row_index":hi,"columns":header,"row_count":count,"sample_rows":samples})
    wb.close(); _CACHE[key]=prof; return prof

def profile_xls(path):
    try: return profile_xlsx(path)
    except Exception: pass
    prof={"path":str(path),"kind":"xls","note":"Legacy .xls not readable with openpyxl"}
    _CACHE[str(path)]=prof; return prof

def profile_file(path):
    if not path.exists(): return {"path":str(path),"error":"missing"}
    ext=path.suffix.lower()
    if ext==".gsheet": return profile_gsheet(path)
    if ext==".csv": return profile_csv(path)
    if ext==".xlsx": return profile_xlsx(path)
    if ext==".xls": return profile_xls(path)
    return {"path":str(path),"kind":ext}

def list_folder_inventory(folder):
    files=sorted([p for p in folder.glob("*") if p.is_file()], key=lambda p:p.name.lower())
    return {"folder":str(folder),"file_count":len(files),"profiled":[profile_file(p) for p in files[:10]],
            "all_filenames":[p.name for p in files]}

def payroll_month_stats(path):
    stats=[]
    if not path.exists(): return stats
    wb=load_workbook(path,read_only=True,data_only=True)
    month_re=re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec|June)-\d{2}$",re.I)
    for name in wb.sheetnames:
        if not month_re.match(name.strip()): continue
        ws=wb[name]; header=None; count=0
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            if i==0: header=[str(c).strip().replace("\\n"," ") if c is not None else "" for c in row]; continue
            if not any(c is not None and str(c).strip() for c in row): continue
            j=next((k for k,h in enumerate(header or []) if "employee code" in h.lower()),None)
            if j is not None and row[j] and str(row[j]).strip().lower() not in {"","asil employee code"}: count+=1
        stats.append({"sheet":name,"employee_rows":count})
    wb.close(); return stats

def fetch_live_counts():
    out={"api_base":LIVE_API,"queried_at":datetime.now(timezone.utc).isoformat()}
    if not JWT_PATH.exists(): out["error"]=f"JWT missing at {JWT_PATH}"; return out
    token=JWT_PATH.read_text(encoding="utf-8").strip(); headers={"Authorization":f"Bearer {token}"}
    def get_json(path):
        req=Request(f"{LIVE_API}{path}",headers=headers)
        with urlopen(req,timeout=120) as resp: return json.loads(resp.read().decode("utf-8"))
    for label,path,key in [("employees","/api/employees","employees"),("contracts","/api/contracts","contracts"),("client_invoices","/api/client-invoices","invoices")]:
        try:
            data=get_json(path); rows=data.get(key,data if isinstance(data,list) else []); out[label]=len(rows)
        except Exception as e: out[f"{label}_error"]=str(e)
    try:
        bills=get_json("/api/bills"); out["bills"]=len(bills if isinstance(bills,list) else bills.get("bills",[]))
    except Exception as e: out["bills_error"]=str(e)
    return out

def summarize_audit_csvs():
    summary={}
    if not AUDIT_OUT.exists(): summary["note"]="audit/out missing"; return summary
    for p in sorted(AUDIT_OUT.glob("*.csv")):
        prof=profile_csv(p); summary[p.name]={"rows":prof.get("row_count"),"columns":prof.get("columns",[])}
    return summary

def build_report():
    report={
      "autodata_monthly":[list_folder_inventory(AUTODATA/f) for f in MONTHLY_FOLDERS if (AUTODATA/f).exists()],
      "april_2026":[profile_file(p) for p in sorted(APRIL_2026.rglob("*")) if p.is_file()] if APRIL_2026.exists() else [],
      "sample_files":[profile_file(p) for p in sorted(SAMPLE_FILES.iterdir()) if p.is_file()] if SAMPLE_FILES.exists() else [],
      "audit_out":summarize_audit_csvs(),
      "payroll_workbook_months":payroll_month_stats(PAYROLL_XLSX),
      "pr_csv_april_2026":profile_csv(PR_CSV) if PR_CSV.exists() else {"error":"missing"},
      "live_hcm":fetch_live_counts(),
    }
    if PAYROLL_XLSX.exists(): report["payroll_workbook_profile"]=profile_xlsx(PAYROLL_XLSX)
    return report

TYPE_MAPPINGS = """
### Payroll month sheet / PR.csv
| Source column | HCM target |
|---|---|
| ASIL Employee Code | employees.id |
| Employee master columns | employees.* via /api/employees/bulk |
| Month sheet Mon-YY | payroll_runs.period_month/year |
| Pay columns | payroll_run_rows.inputs + computed |
| Run metadata | payroll_runs locked, source=excel_import |

### Bank output xlsx
| Beneficiary Account / Amount | employees.bank_account + reconciliation |

### Xero / wafi_invoices_clean.csv
| inv, contact, dates, gross, balance | client_invoices + payment_received_at |

### payments_clean.csv
| amount, pay_date, invoice | invoice_receipts |
"""

GAPS = """
1. AutoData monthly folders are .gsheet stubs only; export xlsx or use audit workbook tabs.
2. Confirm ASIL Employee Code = employees.id.
3. Client/BU to contracts.contract_name mapping required.
4. Workbook missing Feb-25; Nov-25 missing Gratuity; Jan-25 missing PF column.
5. PR.csv Apr-2026: 508 active employees - Build 2 master?
6. Payments: multi-invoice rows, missing inv#, Not Paid/Rejected rows need policy.
7. Wafi scope: 3240 invoices in clean CSV; import all or Jan-25+ only?
8. Pre-2025 PF/gratuity openings not in W6 ledger.
9. WH Income Tax Salaries .xls not readable via openpyxl.
10. April 2026 Invoice Files folder empty.
"""

def write_markdown(report):
    lines=["# HCM Data Migration Map (Build 1)","",datetime.now(timezone.utc).isoformat(),""]
    live=report.get("live_hcm",{})
    lines+=["## Live HCM baseline","","| Entity | Count |","|---|---:|"]
    for k in ("employees","contracts","client_invoices","bills"):
        if k in live: lines.append("| %s | %s |" % (k, live[k]))
    lines+=["","## File inventory",""]
    for inv in report.get("autodata_monthly",[]):
        folder_name=Path(inv["folder"]).name
        lines.append("### AutoData `%s` (%s files)" % (folder_name, inv["file_count"]))
        ext=defaultdict(int)
        for fn in inv.get("all_filenames",[]): ext[Path(fn).suffix.lower() or "(none)"]+=1
        lines+=["","| Ext | Count |","|---|---:|"]
        for e,n in sorted(ext.items()): lines.append("| %s | %s |" % (e,n))
        lines.append("")
    lines+=["### April 2026 outputs","","| File | Kind | Notes |","|---|---|---|"]
    for prof in report.get("april_2026",[]):
        n=Path(prof["path"]).name; note=prof.get("note") or prof.get("error") or ""
        if prof.get("sheets"):
            sh=prof["sheets"][0]; note="%s rows=%s" % (sh["name"], sh.get("row_count"))
        elif prof.get("row_count") is not None: note="rows=%s" % prof["row_count"]
        lines.append("| %s | %s | %s |" % (n, prof.get("kind",""), note))
    lines+=["","### SampleFiles",""]
    for prof in report.get("sample_files",[]):
        n=Path(prof["path"]).name
        if prof.get("sheets"):
            sh=prof["sheets"][0]
            lines.append("- %s: sheet %s cols=%s rows=%s" % (n, sh["name"], len(sh.get("columns",[])), sh.get("row_count")))
        else:
            lines.append("- %s: rows=%s cols=%s" % (n, prof.get("row_count"), len(prof.get("columns",[]))))
    lines+=["","### audit/out","","| File | Rows | Cols |","|---|---:|---:|"]
    for fn,meta in sorted((report.get("audit_out") or {}).items()):
        if fn=="note": continue
        lines.append("| %s | %s | %s |" % (fn, meta.get("rows","?"), len(meta.get("columns",[]))))
    lines+=["","## Payroll month rows","","| Sheet | Employees |","|---|---:|"]
    for st in report.get("payroll_workbook_months",[]):
        lines.append("| %s | %s |" % (st["sheet"], st["employee_rows"]))
    pr=report.get("pr_csv_april_2026",{})
    if pr.get("row_count"):
        lines.append("PR.csv: %s rows, %s columns." % (pr["row_count"], len(pr.get("columns",[]))))
    lines+=["","## Mappings",TYPE_MAPPINGS,"## GAPS & QUESTIONS",GAPS,"## Import order","","| Build | Scope | Est. |","|---|---|---:|"]
    months=report.get("payroll_workbook_months",[]); recent=months[:3] if months else []
    est3=sum(s["employee_rows"] for s in recent); est_all=sum(s["employee_rows"] for s in months)
    wafi=(report.get("audit_out") or {}).get("wafi_invoices_clean.csv",{}).get("rows",3240)
    pays=(report.get("audit_out") or {}).get("payments_clean.csv",{}).get("rows",1400)
    lines.append("| Build 2 | Employee master | ~508 active |")
    lines.append("| Build 3 | %s | ~%s rows (%s all months) |" % (", ".join(s["sheet"] for s in recent), est3, est_all))
    lines.append("| Build 4 | invoices + payments | ~%s inv, ~%s pay |" % (wafi, pays))
    lines+=["","See audit/profile_historic_data_raw.json for sheet-level columns/samples.",""]
    MAP_PATH.parent.mkdir(parents=True,exist_ok=True); MAP_PATH.write_text(chr(10).join(lines),encoding="utf-8")

def main():
    report=build_report()
    RAW_JSON.parent.mkdir(parents=True,exist_ok=True)
    RAW_JSON.write_text(json.dumps(report,indent=2,default=str),encoding="utf-8")
    write_markdown(report)
    print("Profiled %s files; map at %s" % (len(_CACHE), MAP_PATH))

if __name__ == "__main__": main()
