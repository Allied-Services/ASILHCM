#!/usr/bin/env python3
"""Build 3: Import locked historic payroll runs from audit workbook month tabs."""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.error import HTTPError
from urllib.request import Request, urlopen

import openpyxl

REPO = Path(__file__).resolve().parent.parent
# MD Mandate §3 — primary source workbook
WORKBOOK = REPO / "Attachments" / "BPO FM Payroll & Invoice File.xlsx"
# Fallback if primary missing (audit copy)
WORKBOOK_FALLBACK = REPO / "Attachments" / "Audit BPO FM Jul 2026" / "BPO FM Payroll & Invoice File.xlsx"
LIVE_API = "https://asilhcm.onrender.com"
JWT_PATH = Path(r"C:\temp\hcm_jwt.txt")
REPORT_PATH = REPO / "audit" / "payroll_import_report.md"

# ALL historical month tabs present in the master spreadsheet
SHEET_PERIOD = {
    "PR Dec-24": (12, 2024),
    "Jan-25": (1, 2025),
    "Mar-25": (3, 2025),
    "Apr-25": (4, 2025),
    "May-25": (5, 2025),
    "June-25": (6, 2025),
    "Jul-25": (7, 2025),
    "Aug-25": (8, 2025),
    "Sep-25": (9, 2025),
    "Oct-25": (10, 2025),
    "Nov-25": (11, 2025),
    "Dec-25": (12, 2025),
    "Jan-26": (1, 2026),
    "Feb-26": (2, 2026),
    "Mar-26": (3, 2026),
    "Apr-26": (4, 2026),
    "May-26": (5, 2026),
}
# Expected continuous months for gap reporting (Feb-25 missing from workbook)
EXPECTED_PERIODS = [
    (12, 2024), (1, 2025), (2, 2025), (3, 2025), (4, 2025), (5, 2025), (6, 2025),
    (7, 2025), (8, 2025), (9, 2025), (10, 2025), (11, 2025), (12, 2025),
    (1, 2026), (2, 2026), (3, 2026), (4, 2026), (5, 2026),
]

# Fallback when employee has no contractId in HCM
CLIENT_CONTRACT_FALLBACK = {
    "wafi energy pakistan limited": "CTR-1773046722553",
    "wafi energy pakistan pvt ltd": "CTR-1773046722553",
}

NUM = re.compile(r"[^\d.\-]")
CODE_RE = re.compile(r"^ASIL", re.I)


def read_jwt() -> str:
    return JWT_PATH.read_text(encoding="utf-8").strip()


def api_json(method: str, path: str, token: str, body: dict | None = None) -> dict[str, Any]:
    url = f"{LIVE_API}{path}"
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    data = None
    if body is not None:
        data = json.dumps(body).encode("utf-8")
        headers["Content-Type"] = "application/json"
    req = Request(url, data=data, headers=headers, method=method)
    try:
        with urlopen(req, timeout=180) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except HTTPError as e:
        raise RuntimeError(f"{method} {path} HTTP {e.code}: {e.read()[:400]!r}") from e


def num(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if "#NAME?" in s.upper():
        return 0.0
    s = NUM.sub("", s)
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def pick(rec: dict, *keys: str) -> Any:
    for k in keys:
        if k in rec and rec[k] is not None and str(rec[k]).strip() != "":
            return rec[k]
    return None


def row_to_payload(hdr: list[str], row: tuple) -> dict[str, Any] | None:
    rec = {hdr[i]: row[i] if i < len(row) else None for i in range(len(hdr))}
    eid = pick(rec, "ASIL Employee Code")
    if not eid or not CODE_RE.match(str(eid).strip()):
        return None

    gross = num(pick(rec, "Gross Monthly Salary"))
    net = num(pick(rec, "Net Pay for the Month"))
    salary = num(pick(rec, "New Salary"))
    formula_broken = gross <= 0 and salary > 0
    if gross <= 0 and salary > 0:
        gross = salary
    if net <= 0 and gross > 0:
        net = gross
    if gross <= 0 and net <= 0:
        return None

    computed = {
        "salaryForDays": num(pick(rec, "Salary for Days Worked")) or salary,
        "overtimeAmount": num(pick(rec, "Overtime Amount")),
        "gross": gross,
        "wht": num(pick(rec, "Income Tax")),
        "pfDeduction": num(pick(rec, "PF (Deduction)")),
        "eobiEmployee": num(pick(rec, "EOBI (Employee)")),
        "totalDeductions": num(pick(rec, "Total Deductions")),
        "netPay": net,
        "eduCess": num(pick(rec, "Edu. CESS")),
        "sessiEmployer": num(pick(rec, "SESSI")),
        "eobiEmployer": num(pick(rec, "EOBI\n(Employer)", "EOBI (Employer)")),
        "bonusAccrual": num(pick(rec, "Bonus Amount")),
        "gratuityAccrual": num(pick(rec, "Gratuity")),
        "lifeInsurance": num(pick(rec, "Life Insurance  Employee Only")),
        "totalPayrollCost": num(pick(rec, "Total Payroll Cost")) or gross,
        "serviceCharges": num(pick(rec, "Service Charges")),
        "salesTax": num(pick(rec, "Sales Tax")),
        "totalCost": num(pick(rec, "Total Cost")) or gross,
        "billAmount": num(pick(rec, "Total Cost")) or gross,
        "billSource": "excel_import",
    }
    return {
        "employeeId": str(eid).strip(),
        "client": str(pick(rec, "Client") or "").strip(),
        "paidDays": num(pick(rec, "Paid Days")) or None,
        "workingDays": num(pick(rec, "Working Days")) or None,
        "ot2Hours": num(pick(rec, "OT Hrs @ 2X")),
        "ot3Hours": num(pick(rec, "OT Hrs @ 3X")),
        "inputs": {"source": "excel_import", "formula_broken": formula_broken},
        "computed": computed,
    }


def load_employees(token: str) -> dict[str, dict[str, Any]]:
    return {e["id"]: e for e in api_json("GET", "/api/employees", token).get("employees", [])}


def load_contracts(token: str) -> dict[str, str]:
    """Map normalized client / contract_name -> contract id."""
    out: dict[str, str] = {}
    for ct in api_json("GET", "/api/contracts", token).get("contracts", []):
        cid = ct.get("id")
        if not cid:
            continue
        for key in (ct.get("contractName"), ct.get("clientName"), ct.get("client")):
            if key:
                out[str(key).lower().strip()] = cid
    return out


def resolve_contract(
    row: dict[str, Any],
    emp_by_id: dict[str, dict[str, Any]],
    client_to_contract: dict[str, str],
) -> str | None:
    emp = emp_by_id.get(row["employeeId"])
    if emp and emp.get("contractId"):
        return emp["contractId"]
    client_key = row["client"].lower().strip()
    if client_key in client_to_contract:
        return client_to_contract[client_key]
    return CLIENT_CONTRACT_FALLBACK.get(client_key)


def load_sheet(wb, sheet_name: str, emp_by_id: dict[str, dict[str, Any]], client_to_contract: dict[str, str]) -> tuple[dict[str, list[dict]], dict[str, int]]:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
    by_contract: dict[str, list[dict]] = defaultdict(list)
    stats = {"parsed": 0, "importable": 0, "no_contract": 0, "formula_broken": 0}
    for row in rows[1:]:
        payload = row_to_payload(hdr, row)
        if not payload:
            continue
        stats["parsed"] += 1
        if payload["inputs"].get("formula_broken"):
            stats["formula_broken"] += 1
        cid = resolve_contract(payload, emp_by_id, client_to_contract)
        if not cid:
            stats["no_contract"] += 1
            continue
        by_contract[cid].append(payload)
        stats["importable"] += 1
    return dict(by_contract), stats


def write_report(results: list[dict[str, Any]], dry_run: bool, workbook: Path, missing_periods: list[tuple[int, int]]) -> None:
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    try:
        wb_rel = str(workbook.relative_to(REPO))
    except ValueError:
        wb_rel = str(workbook)
    lines = [
        "# Payroll history import (MD Mandate §3 / Build 3)",
        "",
        f"Generated: {now}",
        f"Workbook: `{wb_rel}`",
        f"Mode: {'DRY RUN' if dry_run else 'LIVE'}",
        "",
        "## Missing periods (gap analysis)",
        "",
    ]
    if missing_periods:
        for m, y in missing_periods:
            lines.append(f"- **{m:02d}/{y}** — no sheet in master workbook")
    else:
        lines.append("- None — all expected periods present")
    lines += ["", "## Summary", ""]
    for block in results:
        if "sheet" in block and "stats" in block:
            s = block["stats"]
            lines += [
                f"### {block['sheet']} ({block['month']}/{block['year']})",
                f"- Parsed: {s['parsed']}",
                f"- Importable: {s['importable']}",
                f"- No contract map: {s['no_contract']}",
                f"- Formula fallback: {s['formula_broken']}",
                f"- Contracts: {block.get('contracts', 0)}",
                "",
            ]
        elif block.get("dry_run"):
            lines.append(f"- `{block['contractId']}`: would import {block['rows']} rows ({block['sheet']})")
        elif block.get("ok"):
            lines.append(
                f"- `{block['contractId']}` ({block['sheet']}): run {block.get('runId')} — {block.get('rowsInserted')} rows"
            )
        elif block.get("error"):
            lines.append(f"- `{block.get('contractId')}` ({block.get('sheet')}): **FAIL** {block['error']}")
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--sheets", nargs="*", default=list(SHEET_PERIOD.keys()))
    ap.add_argument("--workbook", default=None)
    args = ap.parse_args()

    workbook = Path(args.workbook) if args.workbook else WORKBOOK
    if not workbook.exists():
        workbook = WORKBOOK_FALLBACK
    if not workbook.exists():
        print(f"Missing workbook: {WORKBOOK} (and fallback)", file=sys.stderr)
        return 1

    present = set(SHEET_PERIOD.values())
    missing_periods = [p for p in EXPECTED_PERIODS if p not in present]
    if missing_periods:
        print("Missing periods in workbook:", missing_periods, flush=True)

    token = read_jwt()
    emp_by_id = load_employees(token)
    client_to_contract = load_contracts(token)
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    results: list[dict[str, Any]] = []

    for sheet in args.sheets:
        if sheet not in SHEET_PERIOD:
            print(f"Skip unknown sheet: {sheet}", file=sys.stderr)
            continue
        if sheet not in wb.sheetnames:
            print(f"Sheet absent in workbook: {sheet}", file=sys.stderr)
            results.append({"sheet": sheet, "ok": False, "error": "sheet_absent"})
            continue
        month, year = SHEET_PERIOD[sheet]
        by_contract, stats = load_sheet(wb, sheet, emp_by_id, client_to_contract)
        results.append({"sheet": sheet, "month": month, "year": year, "stats": stats, "contracts": len(by_contract)})
        print(f"{sheet}: {stats['importable']} rows across {len(by_contract)} contracts", flush=True)

        for contract_id, rows in sorted(by_contract.items()):
            body = {"contractId": contract_id, "month": month, "year": year, "rows": rows}
            if args.dry_run:
                results.append({"sheet": sheet, "contractId": contract_id, "rows": len(rows), "dry_run": True})
                continue
            try:
                res = api_json("POST", "/api/admin/import-payroll-history", token, body)
                results.append({"sheet": sheet, "contractId": contract_id, **res})
            except RuntimeError as err:
                results.append({"sheet": sheet, "contractId": contract_id, "ok": False, "error": str(err)})
            time.sleep(0.25)

    wb.close()
    write_report(results, args.dry_run, workbook, missing_periods)
    print(f"Report: {REPORT_PATH}")
    fails = sum(1 for r in results if r.get("ok") is False)
    return 1 if fails else 0


if __name__ == "__main__":
    raise SystemExit(main())
