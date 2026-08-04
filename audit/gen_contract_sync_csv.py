"""Generate partial CSV for payroll team: Contract ID + Active sync only."""
import csv, json, os, re, psycopg2
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(r'G:\My Drive\Experiments\BPOFMSystem')
ENV = ROOT / 'backend' / '.env'
for line in ENV.read_text().splitlines():
    m = re.match(r'^([^#=]+)=(.*)$', line.strip())
    if m: os.environ.setdefault(m.group(1), m.group(2).strip().strip('"'))

wb = load_workbook(ROOT / 'Attachments' / 'Master Data Updation - July 26.xlsx', read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c else '' for c in rows[0]]
idx = {h: i for i, h in enumerate(hdr)}

conn = psycopg2.connect(os.environ['DATABASE_URL'])
cur = conn.cursor()
cur.execute('SELECT id, contract_id, active FROM employees')
hcm = {r[0]: {'contract_id': r[1] or '', 'active': r[2] or ''} for r in cur.fetchall()}
conn.close()

out_rows = []
for r in rows[1:]:
    if not r or not r[idx.get('ASIL Employee Code', 0)]:
        continue
    eid = str(r[idx['ASIL Employee Code']]).strip()
    if eid not in hcm:
        continue
    ex_cid = str(r[idx.get('Contract ID', '')] or '').strip()
    ex_active = str(r[idx.get('Active', '')] or '').strip()
    ex_active_norm = ex_active.lower()
    if ex_active_norm in ('yes', 'true', '1', 'active'):
        ex_active = 'Yes'
    elif ex_active_norm in ('no', 'false', '0', 'inactive'):
        ex_active = 'No'
    ex_cname = str(r[idx.get('Contract Name', '')] or '').strip()
    if not ex_cid:
        continue
    emp = hcm[eid]
    if ex_cid != emp['contract_id']:
        out_rows.append({
            'ASIL Employee Code': eid,
            'Contract ID': ex_cid,
            'Contract Name': ex_cname,
            'Active': ex_active,
        })

out_path = ROOT / 'audit' / 'payroll_team_contract_sync_july.csv'
with out_path.open('w', encoding='utf-8', newline='') as f:
    w = csv.DictWriter(f, fieldnames=['ASIL Employee Code', 'Contract ID', 'Contract Name', 'Active'])
    w.writeheader()
    w.writerows(out_rows)

print(f'Wrote {len(out_rows)} rows to {out_path}')
